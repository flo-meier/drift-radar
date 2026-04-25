#!/usr/bin/env python3
"""Export Drift Radar data into downloadable artifacts under drift-radar-app/public/downloads/.

Produces:
  - drift_radar.json       (full pipeline output, already exists)
  - drift_radar.csv        (flat CSV)
  - content_briefs/<slug>.pdf  (one Content Brief PDF per Deep-Dive)
  - executive_summary.pdf  (one-page management summary)
"""
import csv
import json
import re
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
)

ROOT = Path(__file__).parent
UI_JSON = ROOT / "data" / "ui" / "drift_radar.json"
OUT = ROOT.parent / "public" / "downloads"
BRIEFS = OUT / "content_briefs"

INK = colors.HexColor("#1A1614")
INK_SOFT = colors.HexColor("#4A413B")
INK_MUTE = colors.HexColor("#8A7F75")
RULE = colors.HexColor("#D4CCBF")
DRIFT = colors.HexColor("#CC7A00")


def slugify(text, max_len=48):
    # ASCII-only to match JavaScript's \w behaviour; umlauts/accents are dropped.
    s = re.sub(r"[^\w\s-]", "", text.lower(), flags=re.ASCII).strip()
    s = re.sub(r"[\s_-]+", "-", s)
    return s[:max_len].rstrip("-")


def brief_filename(prompt, kind=None):
    """Unique filename combining prompt-id suffix and slug, avoiding collisions."""
    pid = prompt["prompt_id"].split("pr_", 1)[-1].split("-", 1)[0][:8]
    prefix = {"own_only": "own-only", "full_silence": "full-silence"}.get(kind)
    slug = slugify(prompt["prompt_text"], max_len=44)
    if prefix:
        return f"brief-{prefix}-{pid}-{slug}.pdf"
    return f"brief-{pid}-{slug}.pdf"


def styles():
    ss = getSampleStyleSheet()
    return {
        "hero": ParagraphStyle(
            "hero", parent=ss["Title"], fontName="Times-Bold", fontSize=26,
            leading=30, textColor=INK, spaceAfter=6, alignment=TA_LEFT,
        ),
        "kicker": ParagraphStyle(
            "kicker", parent=ss["Normal"], fontName="Helvetica-Bold",
            fontSize=9, leading=12, textColor=INK_MUTE, spaceAfter=4,
        ),
        "dek": ParagraphStyle(
            "dek", parent=ss["Normal"], fontName="Times-Italic",
            fontSize=13, leading=17, textColor=INK_SOFT, spaceAfter=14,
        ),
        "h2": ParagraphStyle(
            "h2", parent=ss["Heading2"], fontName="Times-Bold",
            fontSize=15, leading=20, textColor=INK, spaceBefore=12, spaceAfter=6,
        ),
        "h3": ParagraphStyle(
            "h3", parent=ss["Heading3"], fontName="Times-Bold",
            fontSize=12, leading=16, textColor=INK, spaceBefore=6, spaceAfter=4,
        ),
        "body": ParagraphStyle(
            "body", parent=ss["Normal"], fontName="Helvetica",
            fontSize=10, leading=14, textColor=INK_SOFT, spaceAfter=8,
        ),
        "body_ink": ParagraphStyle(
            "body_ink", parent=ss["Normal"], fontName="Helvetica",
            fontSize=10, leading=14, textColor=INK, spaceAfter=8,
        ),
        "small": ParagraphStyle(
            "small", parent=ss["Normal"], fontName="Helvetica",
            fontSize=8, leading=10, textColor=INK_MUTE,
        ),
        "drift_number": ParagraphStyle(
            "drift_number", parent=ss["Normal"], fontName="Courier-Bold",
            fontSize=40, leading=44, textColor=DRIFT, spaceAfter=2,
        ),
        "tldr": ParagraphStyle(
            "tldr", parent=ss["Normal"], fontName="Helvetica",
            fontSize=10.5, leading=15, textColor=INK,
            backColor=colors.HexColor("#F4F4F2"),
            borderColor=colors.HexColor("#B5B2AB"), borderWidth=0.5, borderPadding=10,
            spaceBefore=4, spaceAfter=4,
        ),
    }


def rule_line(width_mm=170, color=RULE, weight=0.6):
    t = Table([[""]], colWidths=[width_mm * mm], rowHeights=[0.6])
    t.setStyle(TableStyle([("LINEBELOW", (0, 0), (-1, -1), weight, color)]))
    return t


def strong_rule(width_mm=170):
    return rule_line(width_mm, INK, 1.2)


def page_footer(canvas, doc, text_left, text_right):
    canvas.saveState()
    canvas.setFont("Helvetica", 8)
    canvas.setFillColor(INK_MUTE)
    canvas.drawString(20 * mm, 12 * mm, text_left)
    canvas.drawRightString(190 * mm, 12 * mm, text_right)
    canvas.restoreState()


def clean_text(s):
    if s is None:
        return ""
    return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _build_brand_matrix(data):
    """Load raw all_brands_report.json and return {prompt_id: {brand_id: {model_id: visibility}}}.
    Empty dict if raw file missing."""
    raw_path = ROOT / "data" / "raw" / "all_brands_report.json"
    if not raw_path.exists():
        return {}
    raw = json.loads(raw_path.read_text(encoding="utf-8"))
    cols = raw.get("columns", [])
    rows = raw.get("rows", [])
    try:
        pid_i = cols.index("prompt_id")
        mid_i = cols.index("model_id")
        bid_i = cols.index("brand_id")
        viz_i = cols.index("visibility")
    except ValueError:
        return {}
    matrix = {}
    for r in rows:
        pid, mid, bid, viz = r[pid_i], r[mid_i], r[bid_i], r[viz_i]
        if pid is None or mid is None or bid is None:
            continue
        matrix.setdefault(pid, {}).setdefault(bid, {})[mid] = viz
    return matrix


def export_csv(data):
    OUT.mkdir(parents=True, exist_ok=True)
    out_path = OUT / "drift_radar.csv"
    models = [m["id"] for m in data["active_models"]]
    model_names = {m["id"]: m["name"] for m in data["active_models"]}
    competitors = data.get("competitors") or []
    matrix = _build_brand_matrix(data)

    def brand_slug(name):
        return re.sub(r"[^\w]", "_", name.lower()).strip("_")

    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        header = [
            "prompt_id", "prompt_text", "topic", "volume",
            "divergence_score", "cv_visibility", "range_visibility",
            "mean_visibility", "own_silence", "silence_type",
        ]
        for m in models:
            nm = model_names[m].lower().replace(" ", "_")
            header += [f"visibility_{nm}", f"mentions_{nm}"]
        # Competitor max visibility per brand (useful for quick sort)
        for c in competitors:
            header.append(f"competitor_max_{brand_slug(c['name'])}")
        # Per-competitor × per-engine visibility (full matrix)
        for c in competitors:
            for m in models:
                eng = model_names[m].lower().replace(" ", "_")
                header.append(f"competitor_{brand_slug(c['name'])}_{eng}")
        header += ["top_competitors", "gap_urls_top3"]
        w.writerow(header)

        for p in data["prompts"]:
            row = [
                p["prompt_id"], p["prompt_text"], p["topic"], p.get("volume") or "",
                p["divergence_score"], p["cv_visibility"], p["range_visibility"],
                p["mean_visibility"], p["own_silence"], p.get("silence_type") or "",
            ]
            for m in models:
                row.append(p["visibility_by_model"].get(m, 0))
                row.append(p["mentions_by_model"].get(m, 0))
            pid = p["prompt_id"]
            brand_map = matrix.get(pid, {})
            # max per competitor
            for c in competitors:
                per_eng = brand_map.get(c["id"], {})
                mx = max(per_eng.values()) if per_eng else 0
                row.append(round(mx, 4))
            # full per-competitor × per-engine matrix
            for c in competitors:
                per_eng = brand_map.get(c["id"], {})
                for m in models:
                    row.append(round(per_eng.get(m, 0), 4))
            comps = ", ".join(
                f"{c['brand_name']} ({int(c['max_visibility']*100)}%)"
                for c in (p.get("top_competitors") or [])[:5]
            )
            row.append(comps)
            gap = p.get("gap_urls") or []
            row.append(" | ".join(u.get("url", "") for u in gap[:3]))
            w.writerow(row)

    print(f"Wrote CSV: {out_path.name} ({out_path.stat().st_size} bytes, {len(data['prompts'])} rows)")


def export_content_brief(prompt, data, s):
    BRIEFS.mkdir(parents=True, exist_ok=True)
    fname = brief_filename(prompt, kind=None)
    path = BRIEFS / fname

    story = []
    story.append(Paragraph(
        f"DRIFT RADAR · CONTENT BRIEF · {data['date_range']['start']} → {data['date_range']['end']}",
        s["kicker"]))
    story.append(strong_rule())
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"»{clean_text(prompt['prompt_text'])}«", s["hero"]))
    tags_line = " · ".join(clean_text(t) for t in (prompt.get("tag_names") or []))
    story.append(Paragraph(
        f"Topic: <i>{clean_text(prompt.get('topic',''))}</i> &nbsp;·&nbsp; "
        f"volume {prompt.get('volume_bucket','–')} &nbsp;·&nbsp; "
        + (f"tags <i>{tags_line}</i> &nbsp;·&nbsp; " if tags_line else "")
        + f"prompt ID "
        f"<font name='Courier' size='8'>{prompt['prompt_id']}</font>",
        s["small"]))
    story.append(Spacer(1, 14))

    vis_values = list(prompt["visibility_by_model"].values())
    vis_min = int(min(vis_values)*100) if vis_values else 0
    vis_max = int(max(vis_values)*100) if vis_values else 0
    silence_note = f" Silence type: {prompt['silence_type']}." if prompt.get('silence_type') else ""

    score_tbl = Table(
        [[Paragraph(f"{prompt['divergence_score']:.2f}", s["drift_number"]),
          Paragraph(
            f"<b>Divergence score.</b> On a 0 – 1 scale, this prompt sits at "
            f"{prompt['divergence_score']:.2f}. Own-brand visibility ranges from "
            f"{vis_min} % to {vis_max} % across the three active engines.{silence_note}",
            s["body_ink"])]],
        colWidths=[45 * mm, 125 * mm],
    )
    score_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
    ]))
    story.append(score_tbl)
    story.append(Spacer(1, 12))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    story.append(Paragraph("What the engines say", s["h2"]))
    cs = prompt.get("chat_sample") or {}
    cs_summary = cs.get("claim_summary") or {}
    type_counts = cs_summary.get("type_counts_by_model") or {}

    for m in data["active_models"]:
        mid = m["id"]
        per = (cs.get("by_model") or {}).get(mid) or {}
        claims = per.get("claims") or []
        counts = type_counts.get(mid) or {}
        vis = prompt["visibility_by_model"].get(mid, 0)

        story.append(Paragraph(
            f"<b>{clean_text(m['name'])}</b> · own-brand visibility "
            f"<font name='Courier'>{int(vis*100)} %</font>", s["h3"]))

        if per.get("note"):
            story.append(Paragraph(f"<i>{clean_text(per['note'])}</i>", s["body"]))
        elif not claims:
            story.append(Paragraph(
                "<i>No claims extracted for this engine (no representative sample).</i>",
                s["body"]))
        else:
            brands = per.get("brands_mentioned") or []
            if brands:
                brand_str = ", ".join(
                    f"#{b['position']} {clean_text(b['name'])}" for b in brands)
                story.append(Paragraph(
                    f"<font color='#8A7F75'>Brands cited (tracked):</font> "
                    f"<b>{brand_str}</b>", s["body_ink"]))
            else:
                story.append(Paragraph(
                    "<font color='#8A7F75'>Brands cited (tracked):</font> <i>none</i>",
                    s["body"]))

            type_str = " &nbsp;·&nbsp; ".join(
                f"{n} {t}" for t, n in counts.items() if n > 0)
            if type_str:
                story.append(Paragraph(
                    f"<font color='#8A7F75'>Claim mix:</font> {type_str}", s["body"]))

            for c in claims[:3]:
                story.append(Paragraph(
                    f"<font name='Helvetica-Bold' size='8' color='#8A7F75'>"
                    f"{c['type'].upper()}</font> &nbsp; {clean_text(c['text'])}",
                    s["body"]))

            if per.get("caveat"):
                story.append(Paragraph(
                    f"<font color='#CC7A00'><i>Note: {clean_text(per['caveat'])}</i></font>",
                    s["small"]))
        story.append(Spacer(1, 4))

    story.append(Spacer(1, 6))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    story.append(Paragraph("Stable vs. divergent concepts", s["h2"]))
    shared_sub = cs_summary.get("shared_substances") or []
    shared_crit = cs_summary.get("shared_criteria") or []
    if shared_sub or shared_crit:
        if shared_sub:
            story.append(Paragraph(
                f"<b>Stable substance claims across engines:</b> "
                f"{', '.join(clean_text(x) for x in shared_sub)}",
                s["body_ink"]))
        if shared_crit:
            story.append(Paragraph(
                f"<b>Stable criterion claims across engines:</b> "
                f"{', '.join(clean_text(x) for x in shared_crit)}",
                s["body_ink"]))
    else:
        story.append(Paragraph(
            "<i>No claims overlap across 2+ engines. Every engine answers with its own "
            "concepts – the narrative is fully divergent.</i>", s["body"]))

    brand_claims_by_model = {}
    for m in data["active_models"]:
        mid = m["id"]
        cls = ((cs.get("by_model") or {}).get(mid) or {}).get("claims") or []
        brand_claims_by_model[m["name"]] = [
            c["text"] for c in cls if c["type"] == "brand"]
    if any(brand_claims_by_model.values()):
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "<b>Brand claims per engine (divergent by design):</b>", s["body_ink"]))
        for name, bl in brand_claims_by_model.items():
            if bl:
                story.append(Paragraph(
                    f"<font color='#8A7F75'>{clean_text(name)}:</font> "
                    f"{', '.join(clean_text(b) for b in bl)}", s["body"]))

    story.append(Spacer(1, 10))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    # Search-intent section
    si_story = search_intent_story(prompt, s)
    if si_story:
        for flow in si_story:
            story.append(flow)
        story.append(Spacer(1, 4))
        story.append(rule_line())
        story.append(Spacer(1, 10))

    # URL gap section (skip if no gap URLs attached)
    gap_story = gap_urls_story(prompt, s)
    if gap_story:
        for flow in gap_story:
            story.append(flow)
        story.append(Spacer(1, 4))
        story.append(rule_line())
        story.append(Spacer(1, 10))

    story.append(Paragraph("Concrete moves", s["h2"]))
    moves = []
    if shared_sub or shared_crit:
        anchors = shared_sub + shared_crit
        moves.append(
            f"<b>Anchor content on the stable concepts.</b> Build or update the product page "
            f"around <i>{', '.join(clean_text(x) for x in anchors[:3])}</i> – these are the "
            f"concepts multiple engines already associate with the category. Use them as "
            f"H2/H3 headlines and in the opening paragraph.")
    else:
        moves.append(
            "<b>No stable concepts detected.</b> Decide which narrative your brand wants to "
            "own for this prompt, then seed it across owned (product page), editorial (blog), "
            "and citeable (study/PDF) assets in parallel.")

    if prompt.get("silence_type") == "own_only":
        tc = prompt.get("top_competitors") or []
        names = ", ".join(clean_text(c['brand_name']) for c in tc[:3])
        moves.append(
            f"<b>Own-only silence.</b> {names} already own this prompt. You are not named "
            "on any engine. Priorities: 1. publish a head-to-head comparison page, "
            "2. place product assets on retailer domains cited by the winning engines, "
            "3. add schema.org/Product markup that mirrors competitor-equivalent spec fields.")
    elif prompt.get("silence_type") == "full":
        moves.append(
            "<b>Full silence – free real estate.</b> No tracked brand is named by any engine. "
            "Create a definitive answer page for this prompt before a competitor does.")

    silent_engines = [m["name"] for m in data["active_models"]
                      if prompt["visibility_by_model"].get(m["id"], 0) == 0]
    winning_engines = [m["name"] for m in data["active_models"]
                       if prompt["visibility_by_model"].get(m["id"], 0) > 0]
    if silent_engines and winning_engines:
        moves.append(
            f"<b>Cross-engine lift.</b> You are visible on "
            f"<font color='#3E5C2E'>{', '.join(winning_engines)}</font> but invisible on "
            f"<font color='#CC7A00'>{', '.join(silent_engines)}</font>. Replicate the winning "
            f"engine's source pattern: inspect <i>top_sources</i> in the raw MCP data for "
            f"this prompt, then pursue placements on those domains.")

    for mv in moves:
        story.append(Paragraph("•  " + mv, s["body_ink"]))
        story.append(Spacer(1, 4))

    story.append(Spacer(1, 12))
    story.append(rule_line())
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "Source: Peec AI MCP · dimensions=[prompt_id, model_id] · Haiku 4.5 claim extraction. "
        "Generated by Drift Radar, built for the Peec AI MCP Challenge 2026 · #BuiltWithPeec",
        s["small"]))

    def add_footer(canvas, doc):
        page_footer(canvas, doc,
                    f"Drift Radar · Content Brief · {clean_text(prompt['prompt_text'])[:60]}",
                    f"{data['date_range']['start']} → {data['date_range']['end']} · Page {doc.page}")

    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=18*mm, bottomMargin=22*mm,
        title=f"Drift Radar – Content Brief – {prompt['prompt_text'][:60]}",
        author="Drift Radar · Built for Peec AI MCP Challenge 2026",
    )
    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    print(f"Wrote brief: {path.name} ({path.stat().st_size} bytes)")


SOURCE_LABELS = {
    "OWN": "You", "COMPETITOR": "Competitor", "EDITORIAL": "Editorial",
    "UGC": "UGC", "REFERENCE": "Reference", "INSTITUTIONAL": "Institutional",
    "CORPORATE": "Corporate", "OTHER": "Other",
}
SOURCE_COLORS = {
    "OWN": colors.HexColor("#3E5C2E"),
    "COMPETITOR": colors.HexColor("#8B5A2B"),
    "EDITORIAL": colors.HexColor("#CC7A00"),
    "UGC": colors.HexColor("#C9A845"),
    "REFERENCE": colors.HexColor("#7A6D92"),
    "INSTITUTIONAL": colors.HexColor("#486E8E"),
    "CORPORATE": colors.HexColor("#1A1614"),
    "OTHER": colors.HexColor("#8A7F75"),
}
SOURCE_ORDER = ["OWN", "COMPETITOR", "EDITORIAL", "UGC", "REFERENCE", "INSTITUTIONAL", "CORPORATE", "OTHER"]

URL_CLASS_LABELS = {
    "HOMEPAGE": "Homepage",
    "CATEGORY_PAGE": "Category page",
    "PRODUCT_PAGE": "Product page",
    "LISTICLE": "Listicle",
    "COMPARISON": "Comparison",
    "PROFILE": "Profile",
    "ALTERNATIVE": "Alternatives page",
    "DISCUSSION": "Discussion",
    "HOW_TO_GUIDE": "How-to guide",
    "ARTICLE": "Article",
    "OTHER": "Other",
}


def search_intent_story(prompt, s):
    """Search + shopping sub-queries the engines issued. Content anchors."""
    sq = prompt.get("search_queries") or []
    sh = prompt.get("shopping_queries") or []
    if not sq and not sh:
        return []
    story = [
        Paragraph("Search fan-out · behind this prompt", s["h3"]),
        Paragraph(
            "The actual sub-queries and shopping lookups engines issued to answer this prompt. "
            "Every row is a content-anchor keyword or a product the brand needs to be listed next to.",
            s["body"]),
        Spacer(1, 4),
    ]
    if sq:
        story.append(Paragraph(
            f"<b>Search queries</b> &nbsp;<font color='#8A7F75'>({len(sq)} total)</font>",
            s["body_ink"]))
        for q in sq[:6]:
            story.append(Paragraph(
                f"<font name='Courier' size='9'>• {clean_text(q.get('query',''))}</font>",
                s["small"]))
        story.append(Spacer(1, 6))
    if sh:
        story.append(Paragraph(
            f"<b>Products surfaced</b> &nbsp;<font color='#8A7F75'>({len(sh)} shopping lookups)</font>",
            s["body_ink"]))
        for entry in sh[:3]:
            products = ", ".join(clean_text(p) for p in (entry.get("products") or [])[:6])
            story.append(Paragraph(
                f"<font name='Courier' size='9'>• {clean_text(entry.get('query',''))}</font>",
                s["small"]))
            if products:
                story.append(Paragraph(
                    f"<font size='9'>  → {products}</font>",
                    s["body"]))
    return story


def gap_urls_story(prompt, s):
    """Flowables for the URL-gap section: URLs that cite competitors but not the brand."""
    gap = prompt.get("gap_urls") or []
    if not gap:
        return []
    story = [
        Paragraph("URL gap · cited for competitors, not for you", s["h3"]),
        Paragraph(
            "Source URLs engines retrieved and cited in competitor-named answers for this prompt, "
            "but never next to the brand. Each row is an earned-placement target.",
            s["body"]),
        Spacer(1, 4),
    ]
    for u in gap[:5]:
        cls = u.get("url_classification") or u.get("classification") or "OTHER"
        cls_label = URL_CLASS_LABELS.get(cls, cls.replace("_", " ").title())
        title = (u.get("title") or "").strip() or u.get("url", "")
        cites = u.get("citation_count") or 0
        retr = u.get("retrieval_count") or 0
        story.append(Paragraph(
            f"<b>{clean_text(title[:110])}</b>",
            s["body_ink"]))
        story.append(Paragraph(
            f"<font name='Courier' size='9'>{clean_text(u.get('url',''))}</font>",
            s["small"]))
        story.append(Paragraph(
            f"<font color='#8A7F75'>{cls_label} · cited {cites}× · retrieved {retr}×</font>",
            s["small"]))
        if u.get("excerpt"):
            story.append(Spacer(1, 3))
            story.append(Paragraph(
                f"<i>{clean_text(u['excerpt'])}</i>",
                s["body"]))
            if u.get("angle"):
                story.append(Paragraph(
                    f"<b>Why it ranks.</b> {clean_text(u['angle'])}",
                    s["small"]))
        story.append(Spacer(1, 6))
    return story


def source_mix_story(prompt, s):
    """Returns a list of flowables for the source-mix section of a brief."""
    sm = prompt.get("source_mix")
    if not sm or sm.get("total_retrieved", 0) == 0:
        return [Paragraph(
            "<i>No URL sources were retrieved by the engines for this prompt in the sample window.</i>",
            s["body"])]

    story = []
    by_norm = sm.get("by_class_normalized", {})

    # Distribution list
    dist_items = []
    for cls in SOURCE_ORDER:
        share = by_norm.get(cls) or 0
        if share <= 0:
            continue
        pct = share * 100
        clr = SOURCE_COLORS[cls].hexval()[2:]
        dist_items.append(
            f"<font color='#{clr}'>■</font> "
            f"<b>{SOURCE_LABELS[cls]}</b> "
            f"<font color='#8A7F75'>{pct:.1f} %</font>"
        )
    if dist_items:
        story.append(Paragraph(
            " &nbsp;·&nbsp; ".join(dist_items), s["body_ink"]))

    # Top domains
    top = sm.get("top_domains") or []
    if top:
        story.append(Spacer(1, 6))
        story.append(Paragraph(
            "<b>Top cited domains:</b>", s["body_ink"]))
        for d in top[:6]:
            clr = SOURCE_COLORS.get(d.get("classification") or "OTHER", colors.HexColor("#8A7F75")).hexval()[2:]
            cls_label = SOURCE_LABELS.get(d.get("classification") or "OTHER", "Other")
            story.append(Paragraph(
                f"<font name='Courier' color='#1A1614'>{clean_text(d['domain'])}</font> "
                f"<font color='#{clr}'>— {cls_label}</font> "
                f"<font color='#8A7F75'>retrieved {int((d.get('retrieved_percentage') or 0) * 100)} % of chats</font>",
                s["small"]))
    return story


def export_lite_brief(prompt, data, s, kind):
    """Generate a lite PDF brief for own-only silence or full-silence prompts.

    kind: 'own_only' or 'full_silence'
    """
    BRIEFS.mkdir(parents=True, exist_ok=True)
    fname = brief_filename(prompt, kind=kind)
    path = BRIEFS / fname

    kicker_label = {
        "own_only": "OWN-ONLY SILENCE BRIEF",
        "full_silence": "FULL SILENCE OPPORTUNITY BRIEF",
    }[kind]

    headline_note = {
        "own_only": (
            f"{clean_text(data['own_brand'])} has zero mentions on every tracked engine for "
            "this prompt, while at least one tracked competitor is cited. A direct brand-visibility "
            "gap with a named winner."),
        "full_silence": (
            "No tracked brand is cited by any engine for this prompt. The category is answered "
            "generically – first mover wins."),
    }[kind]

    story = []
    story.append(Paragraph(
        f"DRIFT RADAR · {kicker_label} · {data['date_range']['start']} → {data['date_range']['end']}",
        s["kicker"]))
    story.append(strong_rule())
    story.append(Spacer(1, 8))
    story.append(Paragraph(f"»{clean_text(prompt['prompt_text'])}«", s["hero"]))
    tags_line2 = " · ".join(clean_text(t) for t in (prompt.get("tag_names") or []))
    story.append(Paragraph(
        f"Topic: <i>{clean_text(prompt.get('topic',''))}</i> &nbsp;·&nbsp; "
        f"volume {prompt.get('volume_bucket','–')} &nbsp;·&nbsp; "
        + (f"tags <i>{tags_line2}</i> &nbsp;·&nbsp; " if tags_line2 else "")
        + f"prompt ID "
        f"<font name='Courier' size='8'>{prompt['prompt_id']}</font>",
        s["small"]))
    story.append(Spacer(1, 12))

    # Headline block
    story.append(Paragraph(headline_note, s["dek"]))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    # Engine visibility
    story.append(Paragraph("Engine visibility", s["h3"]))
    vis_items = []
    for m in data["active_models"]:
        v = prompt["visibility_by_model"].get(m["id"], 0)
        vis_items.append(
            f"<b>{clean_text(m['name'])}</b>: "
            f"<font name='Courier'>{int(v * 100)} %</font>")
    story.append(Paragraph(" &nbsp;·&nbsp; ".join(vis_items), s["body_ink"]))
    story.append(Spacer(1, 10))

    # Competitors (own-only only)
    if kind == "own_only":
        tc = prompt.get("top_competitors") or []
        if tc:
            story.append(Paragraph("Who is cited instead", s["h3"]))
            for c in tc[:5]:
                seen_in = ", ".join(
                    f"{clean_text(sv['model'])} {int(sv['visibility'] * 100)} %"
                    for sv in (c.get("seen_in") or [])[:3])
                story.append(Paragraph(
                    f"<b>{clean_text(c['brand_name'])}</b> &nbsp;·&nbsp; "
                    f"top visibility <font name='Courier'>{int(c['max_visibility'] * 100)} %</font> "
                    f"<font color='#8A7F75'>({seen_in})</font>",
                    s["body"]))
            story.append(Spacer(1, 10))

    # Source mix
    story.append(Paragraph("Citation source mix", s["h3"]))
    for flow in source_mix_story(prompt, s):
        story.append(flow)
    story.append(Spacer(1, 12))

    # Search intent
    si_story = search_intent_story(prompt, s)
    if si_story:
        story.append(rule_line())
        story.append(Spacer(1, 10))
        for flow in si_story:
            story.append(flow)
        story.append(Spacer(1, 6))

    # URL gap
    gap_story = gap_urls_story(prompt, s)
    if gap_story:
        story.append(rule_line())
        story.append(Spacer(1, 10))
        for flow in gap_story:
            story.append(flow)
        story.append(Spacer(1, 6))

    story.append(rule_line())
    story.append(Spacer(1, 10))

    # Concrete moves
    story.append(Paragraph("Concrete moves", s["h2"]))
    if kind == "own_only":
        tc = prompt.get("top_competitors") or []
        names = ", ".join(clean_text(c['brand_name']) for c in tc[:3])
        moves = [
            f"<b>Head-to-head comparison page.</b> Build a landing page that positions "
            f"{clean_text(data['own_brand'])} directly against {names} for this exact question. "
            f"Match their spec fields, exceed them on one differentiator.",
            "<b>Retailer placement.</b> Inspect the Top-domain list above: where the winning "
            "engines retrieve product data (Amazon, Masterhorse, Kraemer, Pferdefutter.de), ensure "
            f"{clean_text(data['own_brand'])}'s product assets – images, descriptions, reviews – "
            "are present and schema-tagged.",
            "<b>Schema.org/Product markup.</b> Add structured product data with competitor-"
            "equivalent spec fields so retrievers have the same signals to match against.",
            "<b>Editorial earned mentions.</b> Scan the Editorial rows in the source mix. These "
            "are news/blog domains that already cover the category – pitch a studied-backed angle "
            "to get cited there.",
            "<b>Re-measure in 14 days.</b> Drift Radar will pick up any visibility change on the "
            "next run. A rising visibility above 15 % within 2 weeks indicates the moves are landing.",
        ]
    else:  # full_silence
        moves = [
            "<b>First-mover content page.</b> Category is answered generically – no brand is named. "
            "Publish a definitive answer page under "
            f"<font name='Courier'>{clean_text(data.get('own_brand','')).lower().replace(' ','-')}.de/"
            f"{slugify(prompt['prompt_text'])}</font> that directly answers this question with "
            "structured, citeable content.",
            "<b>Data-backed claim.</b> If possible, publish a small study or aggregated data in the "
            "category – engines reward citable evidence. This lifts you above generic corporate "
            "content.",
            "<b>Editorial outreach.</b> The Editorial share in the source mix shows which news/"
            "blog domains engines trust for this topic area. Pitch a guest piece or interview "
            "there before a competitor does.",
            "<b>UGC presence.</b> If the UGC share is above 5 %, plant the brand into relevant "
            "Reddit / forum threads (disclosed) to seed retrieval signal.",
            "<b>Re-measure in 30 days.</b> Opportunity windows close fast in AI search – competitors "
            "will find the same silence. Track if own-visibility moves above 15 % within a month.",
        ]

    for mv in moves:
        story.append(Paragraph("•  " + mv, s["body_ink"]))
        story.append(Spacer(1, 4))

    story.append(Spacer(1, 12))
    story.append(rule_line())
    story.append(Spacer(1, 6))
    story.append(Paragraph(
        "Source: Peec AI MCP · dimensions=[prompt_id] · domain classification by Peec. "
        "Generated by Drift Radar, built for the Peec AI MCP Challenge 2026 · #BuiltWithPeec",
        s["small"]))

    def add_footer(canvas, doc):
        page_footer(canvas, doc,
                    f"Drift Radar · {kicker_label.title()} · {clean_text(prompt['prompt_text'])[:50]}",
                    f"{data['date_range']['start']} → {data['date_range']['end']} · Page {doc.page}")

    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=18*mm, bottomMargin=22*mm,
        title=f"Drift Radar – {kicker_label.title()} – {prompt['prompt_text'][:60]}",
        author="Drift Radar · Built for Peec AI MCP Challenge 2026",
    )
    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    print(f"Wrote lite brief: {path.name} ({path.stat().st_size} bytes)")


def _vis_bar(value, max_val, width_mm=70, color=DRIFT, track=colors.HexColor("#EDE8DC")):
    """Horizontal bar as a 2-column Table whose column widths encode the value."""
    v = max(0.0, min(1.0, (value / max_val) if max_val > 0 else 0))
    # Ensure both cells have strictly positive width so reportlab doesn't collapse them.
    min_fill = 0.5  # mm, so zero-values still show a sliver of colour
    fill_mm = max(min_fill, width_mm * v)
    rest_mm = max(0.5, width_mm - fill_mm)
    t = Table([["", ""]], colWidths=[fill_mm * mm, rest_mm * mm], rowHeights=[3.5 * mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), color),
        ("BACKGROUND", (1, 0), (1, 0), track),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return t


def _engine_stats(data):
    out = []
    for m in data["active_models"]:
        mid = m["id"]
        vals = [p["visibility_by_model"].get(mid, 0) for p in data["prompts"]]
        nonzero = sum(1 for v in vals if v > 0)
        out.append({
            "id": mid, "name": m["name"],
            "avg": (sum(vals) / len(vals)) if vals else 0,
            "nonzero": nonzero, "total": len(vals),
            "max": max(vals) if vals else 0,
        })
    return out


def export_executive_summary(data, s):
    path = OUT / "executive_summary.pdf"
    story = []
    brand = clean_text(data["own_brand"])
    eng = _engine_stats(data)
    max_avg = max((e["avg"] for e in eng), default=0)
    min_avg = min((e["avg"] for e in eng), default=0)
    spread_ratio = (max_avg / min_avg) if min_avg > 0 else float("inf")
    spread_label = "∞×" if spread_ratio == float("inf") else f"{spread_ratio:.1f}×"
    bar_scale = max(max_avg * 1.15, 0.10)

    story.append(Paragraph(
        f"DRIFT RADAR · EXECUTIVE SUMMARY · {data['date_range']['start']} → {data['date_range']['end']}",
        s["kicker"]))
    story.append(strong_rule())
    story.append(Spacer(1, 8))
    # Forwarding-line TL;DR – the single sentence a CMO can paste into an email subject.
    own_sil = data['summary']['own_silence_count']
    total = data['summary']['total_prompts']
    own_sil_pct = data['summary'].get('own_silence_percent') or round(own_sil / total * 100) if total else 0
    drifting = data['summary']['high_divergence_count']
    tldr_line = (
        f"<b>TL;DR.</b> Across {total} tracked prompts and "
        f"{len(data['active_models'])} AI engines, {brand} is absent on every engine "
        f"in <b>{own_sil}</b> prompts ({own_sil_pct:.0f} %), "
        f"<b>{drifting}</b> prompts drift between engines, and engine visibility spreads "
        f"<b>{spread_label}</b> between strongest and weakest. "
        f"Pages 2 – 3 list the 5 highest-priority risks, the 5 first-mover opportunities, "
        f"and the divergence-score distribution."
    )
    story.append(Paragraph(tldr_line, s["tldr"]))
    story.append(Spacer(1, 12))
    story.append(Paragraph(
        f"Drift Radar – {brand}", s["hero"]))
    story.append(Paragraph(
        f"How three AI engines describe the same brand – differently. "
        f"Dataset: {data['summary']['total_prompts']} prompts, "
        f"{', '.join(m['name'] for m in data['active_models'])}, "
        f"{data['date_range']['start']} → {data['date_range']['end']}.",
        s["dek"]))

    # --- Engine-split hero --------------------------------------------------
    top_engine = max(eng, key=lambda e: e["avg"]) if eng else None
    bot_engine = min(eng, key=lambda e: e["avg"]) if eng else None
    hero_line = (
        f"<b>Same brand. Same 50 questions. Three engine narratives.</b> "
        f"{brand} shows up in "
        f"<font color='#3E5C2E'><b>{top_engine['nonzero']}</b></font> of "
        f"{top_engine['total']} prompts on {clean_text(top_engine['name'])}, "
        f"but only "
        f"<font color='#CC7A00'><b>{bot_engine['nonzero']}</b></font> on "
        f"{clean_text(bot_engine['name'])}. "
        f"Aggregate visibility spreads <b>{spread_label}</b> between the strongest and weakest engine."
    ) if top_engine and bot_engine else ""
    story.append(Paragraph(hero_line, s["body_ink"]))
    story.append(Spacer(1, 10))

    bar_rows = []
    for e in eng:
        label = Paragraph(
            f"<b>{clean_text(e['name'])}</b>", s["body_ink"])
        bar = _vis_bar(e["avg"], bar_scale)
        pct = Paragraph(
            f"<font name='Courier-Bold' color='#1A1614'>{e['avg']*100:.1f} %</font>"
            f"<br/><font name='Helvetica' size='8' color='#8A7F75'>"
            f"non-zero in {e['nonzero']}/{e['total']} prompts</font>",
            s["small"])
        bar_rows.append([label, bar, pct])
    bars = Table(bar_rows, colWidths=[35 * mm, 90 * mm, 45 * mm])
    bars.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, RULE),
    ]))
    story.append(bars)
    story.append(Spacer(1, 14))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    # --- KPI grid (4 headline numbers) -------------------------------------
    story.append(Paragraph("The numbers", s["h2"]))
    kpi_data = [
        [Paragraph(f"{data['summary']['own_silence_percent']:.0f} %", s["drift_number"]),
         Paragraph(
            f"<b>Own-brand silence.</b> {data['summary']['own_silence_count']} of "
            f"{data['summary']['total_prompts']} prompts do not mention "
            f"{brand} on any active engine.",
            s["body_ink"])],
        [Paragraph(str(data['summary']['own_only_silence_count']),
                   ParagraphStyle("n1", parent=s["drift_number"], textColor=INK)),
         Paragraph(
            "<b>Own-only silence.</b> Prompts where at least one tracked competitor is cited "
            "and the brand is not. The category is being answered – by somebody else.",
            s["body_ink"])],
        [Paragraph(str(data['summary']['high_divergence_count']),
                   ParagraphStyle("n2", parent=s["drift_number"], textColor=INK)),
         Paragraph(
            "<b>Drifting prompts.</b> Divergence score ≥ 0.30. Brand visibility differs by at "
            "least 0.30 between the most- and least-favouring engine on the same question.",
            s["body_ink"])],
        [Paragraph(spread_label,
                   ParagraphStyle("n3", parent=s["drift_number"], textColor=DRIFT)),
         Paragraph(
            "<b>Engine spread.</b> Ratio between the strongest and weakest engine's average "
            "own-brand visibility. A brand can be ubiquitous on one engine and invisible on "
            "another – and traditional visibility dashboards aggregate this away.",
            s["body_ink"])],
    ]
    kpi = Table(kpi_data, colWidths=[35 * mm, 135 * mm])
    kpi.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LINEBELOW", (0, 0), (-1, -2), 0.4, RULE),
    ]))
    story.append(kpi)
    story.append(Spacer(1, 10))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    # --- Top 5 risks with mini engine strip --------------------------------
    def _mini_strip(p):
        """Compact three-bar indicator showing per-engine visibility for the row."""
        cells = []
        for m in data["active_models"]:
            v = p["visibility_by_model"].get(m["id"], 0)
            col = colors.HexColor("#3E5C2E") if v >= 0.5 else (
                DRIFT if v == 0 else colors.HexColor("#C9A845"))
            cells.append(Table(
                [[Paragraph(
                    f"<font size='7' color='white'><b>{int(v * 100)}</b></font>"
                    if v > 0 else f"<font size='7' color='white'><b>0</b></font>",
                    ParagraphStyle("mini", fontName="Helvetica-Bold", alignment=1)
                )]],
                colWidths=[11 * mm], rowHeights=[5 * mm],
                style=TableStyle([
                    ("BACKGROUND", (0, 0), (-1, -1), col),
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 0),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 0),
                    ("TOPPADDING", (0, 0), (-1, -1), 0),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
                ])
            ))
        strip = Table([cells], colWidths=[11 * mm] * len(cells))
        strip.setStyle(TableStyle([
            ("LEFTPADDING", (0, 0), (-1, -1), 1),
            ("RIGHTPADDING", (0, 0), (-1, -1), 1),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        return strip

    own_only = [p for p in data["prompts"] if p.get("silence_type") == "own_only"]
    own_only.sort(key=lambda p: (-(p.get("volume") or 0), -p["divergence_score"]))
    story.append(Paragraph("Top 5 risks · own-only silence", s["h2"]))
    story.append(Paragraph(
        "Prompts where competitors are cited and the brand is not, sorted by search volume. "
        "Engine strip shows own-brand visibility on "
        f"{' · '.join(m['name'] for m in data['active_models'])}.",
        s["body"]))
    story.append(Spacer(1, 4))
    risk_rows = []
    for p in own_only[:5]:
        top3 = ", ".join(clean_text(c["brand_name"]) for c in (p.get("top_competitors") or [])[:3])
        txt = Paragraph(
            f"<b>»{clean_text(p['prompt_text'])}«</b><br/>"
            f"<font size='8' color='#8A7F75'>volume {p.get('volume_bucket','–')} &nbsp;·&nbsp; "
            f"cited instead: <font color='#8B5A2B'>{top3}</font></font>",
            s["body_ink"])
        risk_rows.append([txt, _mini_strip(p)])
    risk_tbl = Table(risk_rows, colWidths=[133 * mm, 37 * mm])
    risk_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, RULE),
    ]))
    story.append(risk_tbl)

    story.append(Spacer(1, 10))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    # --- Top 5 opportunities -----------------------------------------------
    full_sil = [p for p in data["prompts"] if p.get("silence_type") == "full"]
    full_sil.sort(key=lambda p: -(p.get("volume") or 0))
    story.append(Paragraph("Top 5 opportunities · full silence", s["h2"]))
    story.append(Paragraph(
        "Prompts no tracked brand is named in. Free real estate – first mover wins.",
        s["body"]))
    story.append(Spacer(1, 4))
    opp_rows = []
    for p in full_sil[:5]:
        txt = Paragraph(
            f"<b>»{clean_text(p['prompt_text'])}«</b><br/>"
            f"<font size='8' color='#8A7F75'>volume {p.get('volume_bucket','–')} &nbsp;·&nbsp; "
            f"topic: <i>{clean_text(p.get('topic',''))}</i></font>",
            s["body_ink"])
        opp_rows.append([txt, _mini_strip(p)])
    opp_tbl = Table(opp_rows, colWidths=[133 * mm, 37 * mm])
    opp_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, RULE),
    ]))
    story.append(opp_tbl)

    story.append(Spacer(1, 14))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    # --- Divergence distribution histogram (KeepTogether) -----------------
    hist_bins = [(0.0, 0.1, "0 – 0.1"), (0.1, 0.2, "0.1 – 0.2"),
                 (0.2, 0.3, "0.2 – 0.3"), (0.3, 0.5, "0.3 – 0.5"),
                 (0.5, 0.7, "0.5 – 0.7"), (0.7, 1.0, "0.7 – 1.0")]
    hist_counts = []
    for lo, hi, _ in hist_bins:
        c = sum(1 for p in data["prompts"]
                if lo <= p.get("divergence_score", 0) < hi
                or (hi == 1.0 and p.get("divergence_score", 0) == 1.0))
        hist_counts.append(c)
    hist_max = max(hist_counts) or 1
    hist_rows = []
    for (lo, hi, label), c in zip(hist_bins, hist_counts):
        is_drifting = lo >= 0.3
        bar_color = DRIFT if is_drifting else INK
        hist_rows.append([
            Paragraph(f"<font name='Courier'>{label}</font>", s["small"]),
            _vis_bar(c, hist_max, width_mm=100, color=bar_color),
            Paragraph(f"<font name='Courier-Bold' color='#1A1614'>{c}</font>", s["small"]),
        ])
    hist_tbl = Table(hist_rows, colWidths=[30 * mm, 110 * mm, 30 * mm])
    hist_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(KeepTogether([
        Paragraph("Divergence score distribution", s["h2"]),
        Paragraph(
            f"How the {data['summary']['total_prompts']} prompts split across the 0 – 1 divergence scale. "
            "The bump on the right is the drifting shortlist – prompts where the engines genuinely disagree.",
            s["body"]),
        Spacer(1, 6),
        hist_tbl,
        Spacer(1, 4),
        Paragraph(
            "<i>Drifting threshold 0.30 is marked in drift-red. Below that, engines broadly agree "
            "– above, the brand narrative is fragmenting across engines.</i>",
            s["small"]),
    ]))

    story.append(Spacer(1, 14))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    # --- Peec-feature-ready block ------------------------------------------
    story.append(Paragraph("Peec-feature readiness", s["h2"]))
    story.append(Paragraph(
        "Drift Radar composes three native Peec MCP tools – <font name='Courier'>get_brand_report</font> "
        "with <font name='Courier'>dimensions=[prompt_id, model_id]</font>, <font name='Courier'>get_domain_report</font> "
        "with the gap filter, and <font name='Courier'>get_chat</font> – into three new aggregates "
        "Peec could ship as first-class tools:",
        s["body"]))
    story.append(Spacer(1, 4))
    tool_rows = [
        [Paragraph("<font name='Courier' size='9'><b>get_drift_report</b></font>", s["body_ink"]),
         Paragraph("Returns divergence score, per-engine visibility, Wilson CI, and silence classification per prompt. "
                   "Replaces a 50-prompt × 3-engine pivot over <font name='Courier'>get_brand_report</font>.",
                   s["small"])],
        [Paragraph("<font name='Courier' size='9'><b>list_silent_prompts</b></font>", s["body_ink"]),
         Paragraph("Returns prompts where own-visibility is zero across all active engines, partitioned "
                   "into own-only (competitors cited) and full (nobody named). Content-brief priority list.",
                   s["small"])],
        [Paragraph("<font name='Courier' size='9'><b>get_cross_engine_brief</b></font>", s["body_ink"]),
         Paragraph("Returns a normalised narrative comparison for one prompt across all engines, with typed "
                   "claim extraction and stable-vs-divergent concept flags. Drop-in for a content manager's brief.",
                   s["small"])],
    ]
    tool_tbl = Table(tool_rows, colWidths=[48 * mm, 122 * mm])
    tool_tbl.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("RIGHTPADDING", (0, 0), (-1, -1), 0),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -2), 0.3, RULE),
    ]))
    story.append(tool_tbl)
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        "Tool manifest, input/output schemas and example payloads: "
        "<font name='Courier'>drift-radar-mcp.json</font> in the Export kit.",
        s["small"]))

    story.append(Spacer(1, 14))
    story.append(rule_line())
    story.append(Spacer(1, 10))

    # --- Methodology + limitations (merged, compact) ---------------------
    method_lines = [
        Paragraph("Methodology &amp; limitations", s["h3"]),
        Paragraph(
            f"<b>Data.</b> Peec AI MCP (OAuth). {data['summary']['total_prompts']} prompts × {len(data['active_models'])} engines "
            f"({', '.join(m['name'] for m in data['active_models'])}), "
            f"{data['date_range']['start']} → {data['date_range']['end']}, one chat per prompt × engine × day.",
            s["body"]),
        Paragraph(
            "<b>Scoring.</b> Divergence = 0.7 × range + 0.3 × min(CV, 2.0)/2.0. Range is max − min own-brand "
            "visibility across engines per prompt; CV the coefficient of variation. Wilson 95 % CI reported per "
            "prompt × engine in <font name='Courier'>drift_radar.json → prompts[].wilson_ci_by_model</font> – "
            "N = 3, so bands are wide. The score is a ranking signal, not a point estimate.",
            s["body"]),
        Paragraph(
            "<b>Claim extraction.</b> Claude Haiku 4.5, typed brand · substance · function · condition · "
            "criterion. Directional narrative signal, not factual validation.",
            s["body"]),
        Paragraph(
            "<b>Caveats.</b> Three-day window cannot distinguish a structural gap from sampling wobble – "
            "re-measure weekly before acting on single-prompt shifts. 50 prompts cover one category (horse "
            "supplement feed); method generalises, ratios do not. Engine set (ChatGPT, Gemini, AI Overview) "
            "is Peec Pro-Plan-bound; the spread ratio moves when the set changes.",
            s["body"]),
        Spacer(1, 10),
        Paragraph(
            "Source: Peec AI MCP · Generated by Drift Radar · "
            "Built for the Peec AI MCP Challenge 2026 · #BuiltWithPeec",
            s["small"]),
    ]
    story.append(KeepTogether(method_lines))

    def add_footer(canvas, doc):
        page_footer(canvas, doc,
                    f"Drift Radar · Executive Summary · {brand}",
                    f"{data['date_range']['start']} → {data['date_range']['end']} · Page {doc.page}")

    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=20*mm, rightMargin=20*mm,
        topMargin=18*mm, bottomMargin=22*mm,
        title=f"Drift Radar – Executive Summary – {data['own_brand']}",
        author="Drift Radar · Built for Peec AI MCP Challenge 2026",
    )
    doc.build(story, onFirstPage=add_footer, onLaterPages=add_footer)
    print(f"Wrote executive_summary.pdf ({path.stat().st_size} bytes)")


def export_subset_csv(data, filename, predicate, description):
    """Write a filtered CSV using the same schema as the master CSV."""
    out_path = OUT / filename
    models = [m["id"] for m in data["active_models"]]
    model_names = {m["id"]: m["name"] for m in data["active_models"]}

    subset = [p for p in data["prompts"] if predicate(p)]
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.writer(f)
        header = ["prompt_id", "prompt_text", "topic", "volume",
                  "divergence_score", "silence_type", "own_silence"]
        for m in models:
            nm = model_names[m].lower().replace(" ", "_")
            header.append(f"visibility_{nm}")
        header += ["top_competitors", "source_classes"]
        w.writerow(header)
        for p in subset:
            row = [p["prompt_id"], p["prompt_text"], p["topic"], p.get("volume") or "",
                   p["divergence_score"], p.get("silence_type") or "", p["own_silence"]]
            for m in models:
                row.append(p["visibility_by_model"].get(m, 0))
            comps = ", ".join(
                f"{c['brand_name']} ({int(c['max_visibility']*100)}%)"
                for c in (p.get("top_competitors") or [])[:5])
            row.append(comps)
            sm = p.get("source_mix")
            if sm and sm.get("by_class_normalized"):
                classes = ", ".join(
                    f"{cls} {int(share*100)}%"
                    for cls, share in sorted(
                        sm["by_class_normalized"].items(),
                        key=lambda kv: -kv[1]))
                row.append(classes)
            else:
                row.append("")
            w.writerow(row)
    print(f"Wrote {filename} ({out_path.stat().st_size} bytes, {len(subset)} rows) – {description}")


def export_xlsx(data):
    """Write drift_radar.xlsx with multi-sheet view for Excel-first users."""
    out_path = OUT / "drift_radar.xlsx"
    models = [m["id"] for m in data["active_models"]]
    model_names = {m["id"]: m["name"] for m in data["active_models"]}

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A1614")
    accent_fill = PatternFill("solid", fgColor="CC7A00")

    # --- Overview sheet -----------------------------------------------------
    ws = wb.active
    ws.title = "Overview"
    summary = data["summary"]
    avgs = {m["id"]: sum(p["visibility_by_model"].get(m["id"], 0)
                          for p in data["prompts"]) / max(len(data["prompts"]), 1)
            for m in data["active_models"]}
    rows = [
        ["Drift Radar – Executive KPIs", ""],
        ["Brand", data["own_brand"]],
        ["Date range", f"{data['date_range']['start']} → {data['date_range']['end']}"],
        ["Total prompts", summary["total_prompts"]],
        ["Active engines", ", ".join(m["name"] for m in data["active_models"])],
        ["", ""],
        ["Own-brand silence", f"{summary['own_silence_count']} / {summary['total_prompts']}  ({summary['own_silence_percent']:.0f} %)"],
        ["Own-only silence", summary["own_only_silence_count"]],
        ["Full silence", summary["full_silence_count"]],
        ["Drifting prompts (divergence ≥ 0.30)", summary["high_divergence_count"]],
        ["Average own visibility", f"{summary['avg_own_visibility']*100:.1f} %"],
        ["", ""],
        ["Per-engine average visibility", ""],
    ]
    for m in data["active_models"]:
        rows.append([m["name"], f"{avgs[m['id']]*100:.1f} %"])
    max_avg, min_avg = max(avgs.values()), min(avgs.values())
    spread = (max_avg / min_avg) if min_avg > 0 else float("inf")
    rows.append(["Engine spread (max / min)", f"{spread:.1f}×" if spread != float('inf') else "∞"])
    for r, row in enumerate(rows, start=1):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            if r == 1:
                cell.font = Font(bold=True, size=14)
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 42

    # --- Read me sheet (column conventions, Wilson CI, format notes) -------
    ws_rm = wb.create_sheet("Read me", 1)
    rm_rows = [
        ["Drift Radar workbook – read this first", ""],
        ["", ""],
        ["Visibility & Wilson columns", "All visibility_*, wilson_lo_*, wilson_hi_*, mean_visibility, range_visibility, "
                                         "and competitor_* columns are 0–1 ratios. Multiply by 100 for percent display, "
                                         "or apply Excel format »0%« to the columns."],
        ["divergence_score", "0.7 × range(visibility) + 0.3 × clamp(CV, 0, 2) / 2. "
                             "Bounded in [0, 1]. ≥ 0.30 is the »drifting« shortlist."],
        ["wilson_lo_<engine> · wilson_hi_<engine>", "Wilson 95 % confidence interval [lower, upper] for "
                                                     "the visibility point estimate, computed over the sampled chat count "
                                                     "(typically N = 3 for a 3-day window). Wide intervals are honest, "
                                                     "not noise – Peec samples once per engine per day."],
        ["silence_type", "own_only = brand absent on every engine, at least one competitor cited; "
                         "full = no tracked brand mentioned on any engine. Empty = active (drifting or aligned)."],
        ["competitor_<brand>", "Wide-format max-visibility per competitor (one column per tracked competitor). "
                                "For pivot work, consider the long-format helper drift_radar.csv or "
                                "drift_radar.json."],
        ["competitor_<brand>_<engine>", "Per-engine competitor visibility for fine-grained pivots. 0 = absent."],
        ["top_competitors", "Pre-formatted comma-separated string of the top 5 cited competitors with peak "
                            "visibility – ready for paste into a deck or report."],
        ["gap_urls_top3", "Pipe-separated URLs that engines cited next to competitors but not next to "
                          "the brand. The earned-placement target list."],
        ["", ""],
        ["Sheet map", ""],
        ["Overview", "Headline KPIs and per-engine averages."],
        ["All prompts", "Master row-per-prompt with full Wilson CI + competitor matrix."],
        ["Drifting", "Filter: divergence_score ≥ 0.30."],
        ["Own-only silence", "Filter: silence_type = own_only."],
        ["Full silence", "Filter: silence_type = full."],
        ["Source mix", "Per-prompt domain-class share of all retrieved citations."],
    ]
    for r, row in enumerate(rm_rows, start=1):
        for c, v in enumerate(row, start=1):
            cell = ws_rm.cell(row=r, column=c, value=v)
            if r == 1:
                cell.font = Font(bold=True, size=14)
            elif c == 1 and v and r > 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws_rm.column_dimensions['A'].width = 38
    ws_rm.column_dimensions['B'].width = 90

    # --- Master prompts sheet ----------------------------------------------
    competitors = data.get("competitors") or []
    matrix = _build_brand_matrix(data)

    def brand_slug(name):
        return re.sub(r"[^\w]", "_", name.lower()).strip("_")

    ws2 = wb.create_sheet("All prompts")
    header = ["prompt_id", "prompt_text", "topic", "volume",
              "divergence_score", "cv_visibility", "range_visibility",
              "mean_visibility", "own_silence", "silence_type"]
    for m in models:
        nm = model_names[m].lower().replace(" ", "_")
        header += [f"visibility_{nm}", f"mentions_{nm}", f"wilson_lo_{nm}", f"wilson_hi_{nm}"]
    for c in competitors:
        header.append(f"competitor_max_{brand_slug(c['name'])}")
    for c in competitors:
        for m in models:
            eng = model_names[m].lower().replace(" ", "_")
            header.append(f"competitor_{brand_slug(c['name'])}_{eng}")
    header += ["top_competitors", "gap_urls_top3"]
    ws2.append(header)
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for p in data["prompts"]:
        row = [p["prompt_id"], p["prompt_text"], p["topic"], p.get("volume"),
               p["divergence_score"], p["cv_visibility"], p["range_visibility"],
               p["mean_visibility"], p["own_silence"], p.get("silence_type") or ""]
        wci = p.get("wilson_ci_by_model") or {}
        for m in models:
            row.append(p["visibility_by_model"].get(m, 0))
            row.append(p["mentions_by_model"].get(m, 0))
            lo, hi = (wci.get(m) or [0, 0])[:2] if wci else (0, 0)
            row.append(lo)
            row.append(hi)
        brand_map = matrix.get(p["prompt_id"], {})
        for c in competitors:
            per_eng = brand_map.get(c["id"], {})
            mx = max(per_eng.values()) if per_eng else 0
            row.append(round(mx, 4))
        for c in competitors:
            per_eng = brand_map.get(c["id"], {})
            for m in models:
                row.append(round(per_eng.get(m, 0), 4))
        row.append(", ".join(
            f"{c['brand_name']} ({int(c['max_visibility']*100)}%)"
            for c in (p.get("top_competitors") or [])[:5]))
        gap = p.get("gap_urls") or []
        row.append(" | ".join(u.get("url", "") for u in gap[:3]))
        ws2.append(row)
    for col in range(1, len(header) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = \
            40 if header[col-1] == "prompt_text" else \
            30 if header[col-1] in ("prompt_id", "topic", "top_competitors", "gap_urls_top3") else 14

    # --- Filtered sheets ----------------------------------------------------
    def add_subset(name, predicate):
        s = wb.create_sheet(name)
        sub_header = ["prompt_id", "prompt_text", "topic", "volume",
                      "divergence_score", "silence_type"]
        for m in models:
            nm = model_names[m].lower().replace(" ", "_")
            sub_header.append(f"visibility_{nm}")
        sub_header += ["top_competitors"]
        s.append(sub_header)
        for cell in s[1]:
            cell.font = header_font
            cell.fill = header_fill
        rows = [p for p in data["prompts"] if predicate(p)]
        for p in rows:
            row = [p["prompt_id"], p["prompt_text"], p["topic"], p.get("volume"),
                   p["divergence_score"], p.get("silence_type") or ""]
            for m in models:
                row.append(p["visibility_by_model"].get(m, 0))
            row.append(", ".join(
                f"{c['brand_name']} ({int(c['max_visibility']*100)}%)"
                for c in (p.get("top_competitors") or [])[:5]))
            s.append(row)
        for col in range(1, len(sub_header) + 1):
            s.column_dimensions[get_column_letter(col)].width = \
                40 if sub_header[col-1] == "prompt_text" else \
                26 if sub_header[col-1] in ("prompt_id", "topic", "top_competitors") else 14
        return len(rows)

    n_drift = add_subset("Drifting", lambda p: p["divergence_score"] >= 0.3)
    n_own = add_subset("Own-only silence", lambda p: p.get("silence_type") == "own_only")
    n_full = add_subset("Full silence", lambda p: p.get("silence_type") == "full")

    # --- Source mix sheet ---------------------------------------------------
    ws_src = wb.create_sheet("Source mix")
    src_header = ["prompt_id", "prompt_text", "total_retrieved"]
    cls_order = ["OWN", "COMPETITOR", "EDITORIAL", "UGC", "REFERENCE", "INSTITUTIONAL", "CORPORATE", "OTHER"]
    src_header += [f"share_{c.lower()}" for c in cls_order]
    ws_src.append(src_header)
    for cell in ws_src[1]:
        cell.font = header_font
        cell.fill = header_fill
    for p in data["prompts"]:
        sm = p.get("source_mix") or {}
        row = [p["prompt_id"], p["prompt_text"], sm.get("total_retrieved", 0)]
        by_norm = sm.get("by_class_normalized") or {}
        for c in cls_order:
            row.append(round(by_norm.get(c, 0), 4))
        ws_src.append(row)
    for col in range(1, len(src_header) + 1):
        ws_src.column_dimensions[get_column_letter(col)].width = \
            40 if src_header[col-1] == "prompt_text" else 18

    wb.save(out_path)
    print(f"Wrote XLSX: {out_path.name} ({out_path.stat().st_size} bytes, 5 sheets, drift={n_drift}, own-only={n_own}, full={n_full})")


def _brief_md(prompt, data, kind="premium"):
    """Return a markdown string mirroring the PDF brief content."""
    lines = []
    lines.append(f"# Drift Radar – »{prompt['prompt_text']}«\n")
    tags_md = ", ".join(prompt.get("tag_names") or [])
    lines.append(
        f"**Prompt ID:** `{prompt['prompt_id']}`  ·  "
        f"**Topic:** {prompt.get('topic','–')}  ·  "
        f"**Volume:** {prompt.get('volume_bucket','–')} (Peec bucket)  ·  "
        + (f"**Tags:** {tags_md}  ·  " if tags_md else "")
        + f"**Date range:** {data['date_range']['start']} → {data['date_range']['end']}\n")

    vis_items = []
    for m in data["active_models"]:
        v = prompt["visibility_by_model"].get(m["id"], 0)
        vis_items.append(f"- **{m['name']}:** {int(v*100)} %")
    lines.append("## Engine visibility\n")
    lines.extend(vis_items)
    lines.append("")

    lines.append(
        f"**Divergence score:** {prompt['divergence_score']:.2f}  "
        f"(range 0 – 1; 0 = engines agree, 1 = total divergence)\n")
    if prompt.get("silence_type"):
        lines.append(f"**Silence type:** `{prompt['silence_type']}`\n")

    if kind == "premium":
        cs = prompt.get("chat_sample") or {}
        cs_sum = cs.get("claim_summary") or {}
        lines.append("## What the engines say\n")
        for m in data["active_models"]:
            mid = m["id"]
            per = (cs.get("by_model") or {}).get(mid) or {}
            claims = per.get("claims") or []
            vis = prompt["visibility_by_model"].get(mid, 0)
            lines.append(f"### {m['name']} – visibility {int(vis*100)} %\n")
            if per.get("note"):
                lines.append(f"_{per['note']}_\n")
                continue
            if not claims:
                lines.append("_No claims extracted (no representative sample)._\n")
                continue
            brands = per.get("brands_mentioned") or []
            if brands:
                lines.append("**Brands cited (tracked):** " +
                             ", ".join(f"#{b['position']} {b['name']}" for b in brands) + "\n")
            else:
                lines.append("**Brands cited (tracked):** _none_\n")
            for c in claims[:4]:
                lines.append(f"- _{c['type']}_ – {c['text']}")
            if per.get("caveat"):
                lines.append(f"\n> {per['caveat']}")
            lines.append("")

        shared_sub = cs_sum.get("shared_substances") or []
        shared_crit = cs_sum.get("shared_criteria") or []
        if shared_sub or shared_crit:
            lines.append("## Stable vs. divergent concepts\n")
            if shared_sub:
                lines.append(f"**Stable substance claims across engines:** {', '.join(shared_sub)}\n")
            if shared_crit:
                lines.append(f"**Stable criterion claims across engines:** {', '.join(shared_crit)}\n")
        else:
            lines.append("## Stable vs. divergent concepts\n")
            lines.append(
                "_No claims overlap across two or more engines. Every engine answers with its own "
                "concepts – the narrative is fully divergent._\n")

    # Competitors for own-only
    if kind == "own_only":
        tc = prompt.get("top_competitors") or []
        if tc:
            lines.append("## Who is cited instead\n")
            for c in tc[:5]:
                seen = ", ".join(f"{sv['model']} {int(sv['visibility']*100)} %"
                                  for sv in (c.get("seen_in") or [])[:3])
                lines.append(
                    f"- **{c['brand_name']}** – top visibility {int(c['max_visibility']*100)} %  ({seen})")
            lines.append("")

    # Source mix
    sm = prompt.get("source_mix") or {}
    if sm.get("total_retrieved", 0) > 0:
        lines.append("## Citation source mix\n")
        by_norm = sm.get("by_class_normalized") or {}
        for cls in ["OWN", "COMPETITOR", "EDITORIAL", "UGC", "REFERENCE", "INSTITUTIONAL", "CORPORATE", "OTHER"]:
            share = by_norm.get(cls)
            if share and share > 0:
                lines.append(f"- **{cls.title()}:** {share*100:.1f} %")
        lines.append("")
        top = sm.get("top_domains") or []
        if top:
            lines.append("**Top cited domains**\n")
            for d in top[:6]:
                pct = int((d.get("retrieved_percentage") or 0) * 100)
                lines.append(
                    f"- `{d['domain']}` – {d.get('classification','OTHER')}, "
                    f"retrieved {pct} % of chats")
            lines.append("")

    # Search fan-out
    sq = prompt.get("search_queries") or []
    sh = prompt.get("shopping_queries") or []
    if sq or sh:
        lines.append("## Search fan-out · behind this prompt\n")
        lines.append("The actual sub-queries and shopping lookups engines issued to answer this prompt. Every row is a content-anchor keyword or a product the brand needs to be listed next to.\n")
        if sq:
            lines.append(f"**Search queries** ({len(sq)} total)\n")
            for q in sq[:6]:
                lines.append(f"- `{q.get('query','')}`")
            lines.append("")
        if sh:
            lines.append(f"**Products surfaced** ({len(sh)} shopping lookups)\n")
            for e in sh[:3]:
                products = ", ".join((e.get("products") or [])[:6])
                lines.append(f"- _{e.get('query','')}_ → {products}")
            lines.append("")

    # URL gap (cited for competitors, not for you)
    gap = prompt.get("gap_urls") or []
    if gap:
        lines.append("## URL gap · cited for competitors, not for you\n")
        lines.append("Source URLs engines retrieved and cited next to competitor answers, but never next to the brand. Each row is an earned-placement target.\n")
        for u in gap[:5]:
            cls = u.get("url_classification") or u.get("classification") or "OTHER"
            cls_label = URL_CLASS_LABELS.get(cls, cls.replace("_"," ").title())
            title = (u.get("title") or u.get("url", "")).strip()
            lines.append(f"- **{title[:110]}** – _{cls_label}_, cited {u.get('citation_count',0)}×, retrieved {u.get('retrieval_count',0)}×  ")
            lines.append(f"  `{u.get('url','')}`")
            if u.get("excerpt"):
                lines.append(f"  > {u['excerpt']}")
                if u.get("angle"):
                    lines.append(f"  **Why it ranks.** {u['angle']}")
        lines.append("")

    lines.append("## Concrete moves\n")
    if kind == "premium":
        cs_sum = (prompt.get("chat_sample") or {}).get("claim_summary") or {}
        anchors = (cs_sum.get("shared_substances") or []) + (cs_sum.get("shared_criteria") or [])
        if anchors:
            lines.append(f"- **Anchor content on the stable concepts.** Build or update the page around "
                         f"_{', '.join(anchors[:3])}_ – use them as H2/H3 headlines and in the opening paragraph.")
        else:
            lines.append("- **No stable concepts detected.** Decide which narrative to own, seed across "
                         "owned (product page), editorial (blog) and citeable (study/PDF) assets in parallel.")
        if prompt.get("silence_type") == "own_only":
            tc = prompt.get("top_competitors") or []
            names = ", ".join(c['brand_name'] for c in tc[:3])
            lines.append(f"- **Own-only silence.** {names} already own this prompt. Priorities: "
                         "(1) head-to-head page, (2) retailer placements on cited domains, "
                         "(3) schema.org/Product markup matching competitor spec fields.")
        elif prompt.get("silence_type") == "full":
            lines.append("- **Full silence – free real estate.** Create a definitive answer page for this "
                         "prompt before a competitor does.")
    elif kind == "own_only":
        tc = prompt.get("top_competitors") or []
        names = ", ".join(c['brand_name'] for c in tc[:3])
        lines.append(f"- **Head-to-head page.** Position {data['own_brand']} directly against {names} for this prompt.")
        lines.append("- **Retailer placement.** Inspect top-cited domains; ensure product assets and schema tags are present.")
        lines.append("- **Schema.org/Product markup** matching competitor-equivalent spec fields.")
        lines.append("- **Editorial earned mentions.** Pitch study-backed angles to domains in the Editorial source class.")
        lines.append("- **Re-measure in 14 days** to confirm visibility lift.")
    else:  # full_silence
        lines.append(f"- **First-mover content page** under `{data['own_brand'].lower().replace(' ','-')}.de/...` "
                     "answering the question with structured, citeable content.")
        lines.append("- **Data-backed claim** – publish a study or aggregated data in the category.")
        lines.append("- **Editorial outreach** to Editorial-class domains trusted for the topic.")
        lines.append("- **UGC presence** via disclosed Reddit / forum seeding, if UGC share > 5 %.")
        lines.append("- **Re-measure in 30 days** – opportunity windows close fast.")

    lines.append("")
    lines.append("---")
    lines.append(f"Source: Peec AI MCP · Drift Radar · Built for the Peec AI MCP Challenge 2026 · #BuiltWithPeec")
    return "\n".join(lines)


def export_brief_md(prompt, data, kind):
    BRIEFS.mkdir(parents=True, exist_ok=True)
    md_kind = None if kind == "premium" else kind
    fname_pdf = brief_filename(prompt, kind=md_kind)
    fname_md = fname_pdf.rsplit(".", 1)[0] + ".md"
    path = BRIEFS / fname_md
    path.write_text(_brief_md(prompt, data, kind=kind), encoding="utf-8")


def export_rss(data):
    """Drift Radar RSS 2.0 feed – drifting + silent prompts as items."""
    from email.utils import formatdate
    from xml.sax.saxutils import escape as xe
    path = OUT / "drift_radar.rss"
    brand = data["own_brand"]
    date_range = data["date_range"]
    pubdate = formatdate(timeval=None, usegmt=True)
    feed_url = "https://drift-radar.pages.dev"
    items = []
    candidates = [p for p in data["prompts"]
                  if p.get("silence_type") in ("own_only", "full") or p.get("divergence_score", 0) >= 0.3]
    candidates.sort(key=lambda p: -(p.get("divergence_score") or 0))
    for p in candidates[:40]:
        title = f"{p['divergence_score']:.2f} · "
        if p.get("silence_type") == "own_only":
            title += f"Own-only silence · »{p['prompt_text']}«"
        elif p.get("silence_type") == "full":
            title += f"Full silence · »{p['prompt_text']}«"
        else:
            title += f"Drifting · »{p['prompt_text']}«"
        vbm = p.get("visibility_by_model") or {}
        model_lines = "\n".join(
            f"  - {m['name']}: {int(vbm.get(m['id'], 0)*100)}%"
            for m in data["active_models"])
        tc = p.get("top_competitors") or []
        comp_line = ", ".join(
            f"{c['brand_name']} ({int(c['max_visibility']*100)}%)"
            for c in tc[:3]) or "none tracked"
        gap = p.get("gap_urls") or []
        gap_lines = ""
        if gap:
            gap_lines = "\nGap URLs:\n" + "\n".join(
                f"  - {u.get('url','')} ({u.get('classification','OTHER')})" for u in gap[:3])
        description = (
            f"Topic: {p.get('topic','-')}\n"
            f"Volume: {p.get('volume_bucket','-')}\n"
            f"Silence: {p.get('silence_type') or 'none'}\n"
            f"Visibility per engine:\n{model_lines}\n"
            f"Cited instead: {comp_line}"
            f"{gap_lines}"
        )
        link = f"{feed_url}/#deepdive-{p['prompt_id']}"
        guid = f"{feed_url}/prompt/{p['prompt_id']}"
        items.append(
            f"  <item>\n"
            f"    <title>{xe(title)}</title>\n"
            f"    <link>{xe(link)}</link>\n"
            f"    <guid isPermaLink=\"false\">{xe(guid)}</guid>\n"
            f"    <pubDate>{pubdate}</pubDate>\n"
            f"    <description>{xe(description)}</description>\n"
            f"    <category>{xe(p.get('silence_type') or 'drifting')}</category>\n"
            f"  </item>"
        )
    rss = (
        "<?xml version=\"1.0\" encoding=\"UTF-8\"?>\n"
        "<rss version=\"2.0\">\n"
        "<channel>\n"
        f"  <title>Drift Radar · {xe(brand)}</title>\n"
        f"  <link>{feed_url}</link>\n"
        f"  <description>Drifting and silent prompts detected by Drift Radar across ChatGPT, Gemini and Google AI Overview. "
        f"Data window {xe(date_range['start'])} → {xe(date_range['end'])}. Source: Peec AI MCP.</description>\n"
        "  <language>en</language>\n"
        "  <generator>Drift Radar</generator>\n"
        f"  <lastBuildDate>{pubdate}</lastBuildDate>\n"
        f"  <atom:link xmlns:atom=\"http://www.w3.org/2005/Atom\" href=\"{feed_url}/downloads/drift_radar.rss\" rel=\"self\" type=\"application/rss+xml\"/>\n"
        + "\n".join(items) + "\n"
        "</channel>\n"
        "</rss>\n"
    )
    path.write_text(rss, encoding="utf-8")
    print(f"Wrote RSS: {path.name} ({path.stat().st_size} bytes, {len(items)} items)")


def export_raw_zip():
    """Bundle all raw Peec MCP responses + generated UI JSON into a reproducibility ZIP."""
    zip_path = OUT / "reproducible_data.zip"
    raw_dir = ROOT / "data" / "raw"
    ui_json = ROOT / "data" / "ui" / "drift_radar.json"
    readme = (
        "Drift Radar – Reproducible data bundle\n"
        "======================================\n\n"
        "Contents\n"
        "--------\n"
        "- raw/ – raw JSON responses from Peec AI MCP tools\n"
        "    all_brands_report.json     get_brand_report across all brands\n"
        "    pferdegold_brand_report.json  per-prompt × model report for Pferdegold\n"
        "    sources_by_prompt.json     get_domain_report dimension=[prompt_id]\n"
        "    chat_samples.json          get_chat payloads used as Deep-Dive samples\n"
        "    claims.json                Haiku-extracted claims\n"
        "    lookup_tables.json         ID ↔ name maps (brands, topics, models)\n"
        "- ui/drift_radar.json – final pipeline output shown in the dashboard\n\n"
        "Reproduction\n"
        "------------\n"
        "1. Peec MCP calls: 03-arbeit/drift-radar/run.py (OAuth-authenticated MCP)\n"
        "2. Claim extraction: 03-arbeit/drift-radar/extract_claims.py (Claude Haiku 4.5)\n"
        "3. Artifacts: 03-arbeit/drift-radar/export_downloads.py (this file)\n\n"
        "License\n"
        "-------\n"
        "Source data © Peec AI (accessed via authorised OAuth). "
        "Derived artifacts © Drift Radar, released for the Peec AI MCP Challenge 2026 under the "
        "challenge IP terms (Peec receives worldwide, royalty-free license).\n"
    )
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("README.txt", readme)
        if ui_json.exists():
            zf.write(ui_json, arcname="ui/drift_radar.json")
        if raw_dir.exists():
            for p in sorted(raw_dir.glob("*.json")):
                zf.write(p, arcname=f"raw/{p.name}")
    print(f"Wrote reproducible_data.zip ({zip_path.stat().st_size} bytes)")


def main():
    data = json.loads(UI_JSON.read_text(encoding="utf-8"))
    s = styles()

    # Hard exports: master + tab-scoped CSVs + XLSX
    export_csv(data)
    export_subset_csv(data, "heatmap_drift.csv",
                      lambda p: p["divergence_score"] >= 0.3,
                      "Drifting prompts only (≥0.30)")
    export_subset_csv(data, "own_only_silence.csv",
                      lambda p: p.get("silence_type") == "own_only",
                      "Own-only silence prompts")
    export_subset_csv(data, "full_silence.csv",
                      lambda p: p.get("silence_type") == "full",
                      "Full silence prompts")
    export_xlsx(data)

    # Executive summary + briefs (PDF + Markdown)
    export_executive_summary(data, s)

    premium = [p for p in data["prompts"] if p.get("chat_sample")]
    for p in premium:
        export_content_brief(p, data, s)
        export_brief_md(p, data, kind="premium")

    own_only = [p for p in data["prompts"]
                if p.get("silence_type") == "own_only" and not p.get("chat_sample")]
    for p in own_only:
        export_lite_brief(p, data, s, kind="own_only")
        export_brief_md(p, data, kind="own_only")

    full_sil = [p for p in data["prompts"]
                if p.get("silence_type") == "full" and not p.get("chat_sample")]
    for p in full_sil:
        export_lite_brief(p, data, s, kind="full_silence")
        export_brief_md(p, data, kind="full_silence")

    export_rss(data)
    export_raw_zip()

    total_briefs = len(premium) + len(own_only) + len(full_sil)
    print(
        f"\nDone: {total_briefs} briefs (PDF + MD) "
        f"({len(premium)} premium · {len(own_only)} own-only · {len(full_sil)} full-silence) "
        f"+ 1 executive summary + 4 CSVs + 1 XLSX + 1 RSS + 1 reproducibility ZIP.")


if __name__ == "__main__":
    main()
